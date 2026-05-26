"""Generate a multi-sheet Excel report from classified posts and (optionally)
push it to Google Drive via the rclone-stored OAuth token.

Sheets:
    1. Executive Summary    — KPIs + top-5 pains/goals + donut chart
    2. Top Pains            — pain × subreddit cross-tab with color scale
    3. Top Goals            — goal × subreddit cross-tab with color scale
    4. Verbatims            — quote_worthy posts for marketing copy
    5. Knowledge Gaps       — unique gap questions with frequency
    6. Demographics         — age/income/country × pain breakdown
    7. Subreddit Profiles   — per-sub stats: posts, top pain, top goal, quotable %
    8. Trends               — fetched_at timeline (placeholder until weekly data)
    9. Raw Data             — full classified-post export for drill-down

Usage:
    python build_report.py                     # write reports/reddit_research_YYYY-MM-DD.xlsx
    python build_report.py --upload            # also push to gdrive:проекты/reddit_parser/
    python build_report.py --upload --replace  # overwrite existing file in place (preserves URL)
"""
from __future__ import annotations

import argparse
import datetime as dt
import logging
import sqlite3
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import DoughnutChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from config import DB_PATH

logger = logging.getLogger(__name__)

# --- Palette --------------------------------------------------------------

NAVY = "1F3A5F"
TEAL = "2E8B8B"
AMBER = "D89B2D"
RED = "C0392B"
GREEN = "27AE60"
GREY_HEAD = "E8ECEF"
GREY_ALT = "F7F9FA"
WHITE = "FFFFFF"

THIN = Side(style="thin", color="C9CFD3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
H_FILL = PatternFill("solid", fgColor=NAVY)
H_FONT = Font(bold=True, color=WHITE, name="Helvetica Neue", size=11)
BODY_FONT = Font(name="Helvetica Neue", size=10)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


# --- DB helpers -----------------------------------------------------------


def fetch_all(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute(
        """
        SELECT id, subreddit_name, title, selftext, url, author, upvotes,
               num_comments, hotness_score, created_utc,
               pain_category, pain_other, goal_category, goal_other,
               knowledge_gap, emotion, life_stage,
               financial_experience, employment_status, family_status,
               intent, interests,
               trigger_event, buying_signal, urgency, mentioned_brands,
               quote_worthy, classified_at
        FROM posts WHERE classified_at IS NOT NULL
        ORDER BY hotness_score DESC
        """
    ).fetchall()


def fetch_demographics(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    return conn.execute("SELECT * FROM demographics").fetchall()


def fetch_subreddits(conn: sqlite3.Connection) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        "SELECT name, subscribers, description FROM subreddits"
    ).fetchall()
    return {r["name"]: dict(r) for r in rows}


def fetch_comments_classified(conn: sqlite3.Connection) -> list[sqlite3.Row]:
    """Classified comments joined with their parent post + subreddit."""
    return conn.execute(
        """
        SELECT c.id            AS comment_id,
               c.post_id       AS post_id,
               c.body          AS body,
               c.upvotes       AS upvotes,
               c.created_utc   AS created_utc,
               c.depth         AS depth,
               c.pain_category AS pain_category,
               c.goal_category AS goal_category,
               c.knowledge_gap AS knowledge_gap,
               c.emotion       AS emotion,
               c.life_stage    AS life_stage,
               c.financial_experience AS financial_experience,
               c.employment_status AS employment_status,
               c.family_status AS family_status,
               c.intent        AS intent,
               c.interests     AS interests,
               c.trigger_event AS trigger_event,
               c.buying_signal AS buying_signal,
               c.urgency       AS urgency,
               c.mentioned_brands AS mentioned_brands,
               p.subreddit_name AS subreddit_name,
               p.title         AS post_title,
               p.url           AS post_url
          FROM comments c
          JOIN posts p ON p.id = c.post_id
         WHERE c.classified_at IS NOT NULL
           AND c.is_deleted = 0
         ORDER BY c.upvotes DESC
        """
    ).fetchall()


# --- Sheet helpers --------------------------------------------------------


def style_header(ws: Worksheet, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = H_FILL
        cell.font = H_FONT
        cell.alignment = CENTER
        cell.border = BORDER


def autosize(ws: Worksheet, widths: dict[int, int]) -> None:
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def write_row(ws: Worksheet, row: int, values: list[Any], alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=GREY_ALT) if alt else None
    for i, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=i, value=val)
        cell.font = BODY_FONT
        cell.border = BORDER
        if isinstance(val, (int, float)):
            cell.alignment = CENTER
        else:
            cell.alignment = LEFT
        if fill:
            cell.fill = fill


def add_color_scale(ws: Worksheet, cell_range: str, vmax: int) -> None:
    """Absolute-value diverging white→navy color scale."""
    if vmax <= 0:
        return
    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color="FFFFFF",
        mid_type="num", mid_value=max(1, vmax // 2), mid_color="A6CFE5",
        end_type="num", end_value=vmax, end_color=NAVY,
    )
    ws.conditional_formatting.add(cell_range, rule)


# --- Sheet builders -------------------------------------------------------


def sheet_summary(
    wb: Workbook,
    posts: list[sqlite3.Row],
    subs: dict[str, dict[str, Any]],
    comments: list[sqlite3.Row] | None = None,
) -> None:
    ws = wb.create_sheet("Executive Summary", 0)
    autosize(ws, {1: 32, 2: 18, 3: 24, 4: 18})

    # Top banner
    ws.merge_cells("A1:D1")
    ws["A1"] = "Reddit Finance Research — Executive Summary"
    ws["A1"].font = Font(bold=True, color=WHITE, size=18, name="Helvetica Neue")
    ws["A1"].fill = H_FILL
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    classified = [p for p in posts if p["classified_at"]]
    quotable = [p for p in classified if p["quote_worthy"]]
    pains = Counter(p["pain_category"] for p in classified if p["pain_category"] and p["pain_category"] != "none")
    goals = Counter(p["goal_category"] for p in classified if p["goal_category"] and p["goal_category"] != "none")
    stages = Counter(p["life_stage"] for p in classified if p["life_stage"] and p["life_stage"] != "unknown")

    # KPI rows
    n_comments = len(comments) if comments else 0
    n_comments_pain = sum(
        1 for c in (comments or []) if c["pain_category"] and c["pain_category"] != "none"
    )

    ws["A3"] = "Report generated"
    ws["B3"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Posts classified"
    ws["B4"] = len(classified)
    ws["A5"] = "Comments classified"
    ws["B5"] = n_comments
    ws["A6"] = "Quotable verbatims"
    ws["B6"] = len(quotable)
    ws["A7"] = "Subreddits covered"
    ws["B7"] = len(subs)
    ws["A8"] = "Post pains identified"
    ws["B8"] = sum(pains.values())
    ws["A9"] = "Comment pains identified"
    ws["B9"] = n_comments_pain
    ws["A10"] = "Unique knowledge gaps"
    ws["B10"] = len({p["knowledge_gap"] for p in classified if p["knowledge_gap"]})

    for r in range(3, 11):
        ws[f"A{r}"].font = Font(bold=True, name="Helvetica Neue", size=11)
        ws[f"A{r}"].border = BORDER
        ws[f"B{r}"].border = BORDER
        ws[f"B{r}"].font = BODY_FONT

    # Top 5 pains
    ws["A12"] = "TOP 5 PAINS"
    ws["A12"].font = Font(bold=True, color=WHITE, size=12, name="Helvetica Neue")
    ws["A12"].fill = PatternFill("solid", fgColor=RED)
    ws.merge_cells("A12:B12")
    for i, (cat, n) in enumerate(pains.most_common(5)):
        ws[f"A{13+i}"] = cat
        ws[f"B{13+i}"] = n
        ws[f"A{13+i}"].font = BODY_FONT
        ws[f"B{13+i}"].font = BODY_FONT
        ws[f"A{13+i}"].border = BORDER
        ws[f"B{13+i}"].border = BORDER

    # Top 5 goals
    ws["C12"] = "TOP 5 GOALS"
    ws["C12"].font = Font(bold=True, color=WHITE, size=12, name="Helvetica Neue")
    ws["C12"].fill = PatternFill("solid", fgColor=GREEN)
    ws.merge_cells("C12:D12")
    for i, (cat, n) in enumerate(goals.most_common(5)):
        ws[f"C{13+i}"] = cat
        ws[f"D{13+i}"] = n
        ws[f"C{13+i}"].font = BODY_FONT
        ws[f"D{13+i}"].font = BODY_FONT
        ws[f"C{13+i}"].border = BORDER
        ws[f"D{13+i}"].border = BORDER

    # Donut chart for pain distribution
    if pains:
        chart_start = 21
        ws.cell(row=chart_start, column=1, value="pain_category").font = Font(bold=True)
        ws.cell(row=chart_start, column=2, value="count").font = Font(bold=True)
        for i, (cat, n) in enumerate(pains.most_common(), start=1):
            ws.cell(row=chart_start + i, column=1, value=cat)
            ws.cell(row=chart_start + i, column=2, value=n)
        last = chart_start + len(pains)
        chart = DoughnutChart()
        chart.title = "Pain category distribution"
        chart.height = 9
        chart.width = 13
        labels = Reference(ws, min_col=1, min_row=chart_start + 1, max_row=last)
        data = Reference(ws, min_col=2, min_row=chart_start, max_row=last)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "C21")


def sheet_crosstab(
    wb: Workbook,
    title: str,
    posts: list[sqlite3.Row],
    field: str,
) -> None:
    ws = wb.create_sheet(title)
    # Build matrix: rows = subreddit, cols = field values
    field_vals: dict[str, int] = Counter()
    sub_vals: dict[str, int] = Counter()
    cross: dict[tuple[str, str], int] = Counter()
    for p in posts:
        cat = p[field]
        if not cat or cat == "none":
            continue
        sub = p["subreddit_name"]
        field_vals[cat] += 1
        sub_vals[sub] += 1
        cross[(sub, cat)] += 1

    if not field_vals:
        ws["A1"] = f"No '{field}' data classified yet."
        ws["A1"].font = BODY_FONT
        return

    cats = [c for c, _ in field_vals.most_common()]
    subs_sorted = [s for s, _ in sub_vals.most_common()]

    # Header
    ws.cell(row=1, column=1, value="subreddit \\ " + field)
    for i, c in enumerate(cats, start=2):
        ws.cell(row=1, column=i, value=c)
    ws.cell(row=1, column=len(cats) + 2, value="TOTAL")
    style_header(ws, 1, len(cats) + 2)

    # Body rows
    vmax = max(cross.values())
    for r, sub in enumerate(subs_sorted, start=2):
        ws.cell(row=r, column=1, value=sub).font = Font(bold=True, name="Helvetica Neue", size=10)
        ws.cell(row=r, column=1).border = BORDER
        for c_idx, c in enumerate(cats, start=2):
            val = cross.get((sub, c), 0)
            cell = ws.cell(row=r, column=c_idx, value=val if val else None)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        ws.cell(row=r, column=len(cats) + 2, value=sub_vals[sub]).font = Font(bold=True)
        ws.cell(row=r, column=len(cats) + 2).border = BORDER
        ws.cell(row=r, column=len(cats) + 2).alignment = CENTER

    # TOTAL row
    total_row = len(subs_sorted) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = H_FILL
    ws.cell(row=total_row, column=1).border = BORDER
    for c_idx, c in enumerate(cats, start=2):
        ws.cell(row=total_row, column=c_idx, value=field_vals[c]).font = Font(bold=True, color=WHITE)
        ws.cell(row=total_row, column=c_idx).fill = H_FILL
        ws.cell(row=total_row, column=c_idx).border = BORDER
        ws.cell(row=total_row, column=c_idx).alignment = CENTER
    grand = sum(sub_vals.values())
    ws.cell(row=total_row, column=len(cats) + 2, value=grand).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=len(cats) + 2).fill = H_FILL
    ws.cell(row=total_row, column=len(cats) + 2).border = BORDER
    ws.cell(row=total_row, column=len(cats) + 2).alignment = CENTER

    # Color scale on the cross-tab interior (exclude header, exclude TOTAL row/col)
    start_cell = ws.cell(row=2, column=2).coordinate
    end_cell = ws.cell(row=total_row - 1, column=len(cats) + 1).coordinate
    add_color_scale(ws, f"{start_cell}:{end_cell}", vmax)

    widths = {1: 26}
    for i in range(2, len(cats) + 2):
        widths[i] = max(12, len(cats[i - 2]) + 2)
    widths[len(cats) + 2] = 10
    autosize(ws, widths)
    ws.freeze_panes = "B2"


def sheet_verbatims(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Verbatims")
    quotable = [p for p in posts if p["quote_worthy"]]
    headers = [
        "subreddit", "pain", "goal", "emotion", "buying_signal", "trigger_event",
        "life_stage", "experience", "title", "selftext", "upvotes", "url",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    for r, p in enumerate(quotable, start=2):
        row = [
            p["subreddit_name"], p["pain_category"], p["goal_category"],
            p["emotion"], p["buying_signal"], p["trigger_event"],
            p["life_stage"], p["financial_experience"],
            p["title"], (p["selftext"] or "")[:2000],
            p["upvotes"], p["url"],
        ]
        write_row(ws, r, row, alt=(r % 2 == 0))

    autosize(ws, {
        1: 20, 2: 20, 3: 20, 4: 12, 5: 30, 6: 22, 7: 14, 8: 14,
        9: 55, 10: 70, 11: 9, 12: 50,
    })
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def sheet_gaps(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Knowledge Gaps")
    gaps: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "examples": []})
    for p in posts:
        g = p["knowledge_gap"]
        if not g:
            continue
        gaps[g]["count"] += 1
        if len(gaps[g]["examples"]) < 3:
            gaps[g]["examples"].append(p["title"])

    headers = ["count", "knowledge_gap", "pain_category", "sample_titles"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    sorted_gaps = sorted(gaps.items(), key=lambda x: -x[1]["count"])
    pain_by_gap = {p["knowledge_gap"]: p["pain_category"] for p in posts if p["knowledge_gap"]}
    for r, (gap, info) in enumerate(sorted_gaps, start=2):
        write_row(
            ws, r,
            [info["count"], gap, pain_by_gap.get(gap, ""), " | ".join(info["examples"])],
            alt=(r % 2 == 0),
        )
    autosize(ws, {1: 8, 2: 60, 3: 22, 4: 80})
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def sheet_demographics(
    wb: Workbook, posts: list[sqlite3.Row], demos: list[sqlite3.Row]
) -> None:
    ws = wb.create_sheet("Demographics")
    pain_by_post = {p["id"]: p["pain_category"] for p in posts}

    # Age buckets × pain
    bucket = lambda a: (  # noqa: E731
        "18-24" if a is not None and a < 25 else
        "25-34" if a is not None and a < 35 else
        "35-49" if a is not None and a < 50 else
        "50+" if a is not None else "unknown"
    )
    cross: dict[tuple[str, str], int] = Counter()
    for d in demos:
        pain = pain_by_post.get(d["source_id"])
        if not pain or pain == "none":
            continue
        cross[(bucket(d["age"]), pain)] += 1

    ws["A1"] = "Age bucket × pain (from demographics table)"
    ws["A1"].font = Font(bold=True, size=12, name="Helvetica Neue")

    if cross:
        cats = sorted({c for _, c in cross.keys()})
        ages = ["18-24", "25-34", "35-49", "50+", "unknown"]
        for i, c in enumerate(cats, start=2):
            ws.cell(row=3, column=i, value=c)
        ws.cell(row=3, column=1, value="age \\ pain")
        ws.cell(row=3, column=len(cats) + 2, value="TOTAL")
        style_header(ws, 3, len(cats) + 2)
        for r, age in enumerate(ages, start=4):
            ws.cell(row=r, column=1, value=age).font = Font(bold=True)
            row_total = 0
            for c_idx, c in enumerate(cats, start=2):
                val = cross.get((age, c), 0)
                ws.cell(row=r, column=c_idx, value=val if val else None).alignment = CENTER
                row_total += val
            ws.cell(row=r, column=len(cats) + 2, value=row_total or None).font = Font(bold=True)
        max_v = max(cross.values())
        if max_v:
            add_color_scale(
                ws,
                f"B4:{get_column_letter(len(cats)+1)}{3 + len(ages)}",
                max_v,
            )

    # Country counts
    country_pain: dict[tuple[str, str], int] = Counter()
    for d in demos:
        if not d["country"]:
            continue
        pain = pain_by_post.get(d["source_id"]) or "none"
        country_pain[(d["country"], pain)] += 1

    start_r = 10 + 5  # below age block
    ws.cell(row=start_r, column=1, value="Country × pain").font = Font(bold=True, size=12)
    if country_pain:
        cats2 = sorted({c for _, c in country_pain.keys()})
        for i, c in enumerate(cats2, start=2):
            ws.cell(row=start_r + 2, column=i, value=c)
        ws.cell(row=start_r + 2, column=1, value="country")
        style_header(ws, start_r + 2, len(cats2) + 1)
        countries = sorted({k[0] for k in country_pain.keys()})
        for r, country in enumerate(countries, start=start_r + 3):
            ws.cell(row=r, column=1, value=country).font = Font(bold=True)
            for c_idx, c in enumerate(cats2, start=2):
                val = country_pain.get((country, c), 0)
                ws.cell(row=r, column=c_idx, value=val if val else None).alignment = CENTER

    autosize(ws, {1: 20, **{i: 18 for i in range(2, 15)}})


def sheet_subreddit_profiles(
    wb: Workbook, posts: list[sqlite3.Row], subs_meta: dict[str, dict[str, Any]]
) -> None:
    ws = wb.create_sheet("Subreddit Profiles")
    by_sub: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for p in posts:
        by_sub[p["subreddit_name"]].append(p)

    headers = ["subreddit", "subscribers", "posts_classified", "top_pain", "top_goal",
               "quotable_%", "top_emotion", "top_life_stage"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    for r, (sub, plist) in enumerate(sorted(by_sub.items(), key=lambda x: -len(x[1])), start=2):
        pains = Counter(p["pain_category"] for p in plist if p["pain_category"] != "none")
        goals = Counter(p["goal_category"] for p in plist if p["goal_category"] != "none")
        emos = Counter(p["emotion"] for p in plist if p["emotion"] != "neutral")
        stages = Counter(p["life_stage"] for p in plist if p["life_stage"] != "unknown")
        quotable_pct = round(100 * sum(1 for p in plist if p["quote_worthy"]) / len(plist), 1)
        write_row(ws, r, [
            sub,
            subs_meta.get(sub, {}).get("subscribers", 0),
            len(plist),
            pains.most_common(1)[0][0] if pains else "-",
            goals.most_common(1)[0][0] if goals else "-",
            quotable_pct,
            emos.most_common(1)[0][0] if emos else "-",
            stages.most_common(1)[0][0] if stages else "-",
        ], alt=(r % 2 == 0))
    autosize(ws, {1: 24, 2: 14, 3: 18, 4: 22, 5: 22, 6: 12, 7: 16, 8: 16})
    ws.freeze_panes = "A2"


def sheet_trends(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Trends")
    ws["A1"] = "Trends — fetched_at distribution"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A3"] = (
        "Cross-week trends become meaningful after 2+ weekly collection runs. "
        "Right now this sheet shows the timestamp distribution of currently-collected posts."
    )
    ws["A3"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A3:F3")

    by_day: Counter[str] = Counter()
    for p in posts:
        if not p["created_utc"]:
            continue
        day = dt.datetime.fromtimestamp(p["created_utc"]).strftime("%Y-%m-%d")
        by_day[day] += 1
    ws["A5"] = "post_created_date"
    ws["B5"] = "posts"
    style_header(ws, 5, 2)
    for i, (day, n) in enumerate(sorted(by_day.items()), start=6):
        ws.cell(row=i, column=1, value=day)
        ws.cell(row=i, column=2, value=n).alignment = CENTER
    autosize(ws, {1: 20, 2: 10})


def sheet_top_comments(wb: Workbook, comments: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Top Comments")
    # Take comments with a real pain/goal signal, ordered by upvotes desc
    rated = [c for c in comments if c["pain_category"] != "none" or c["goal_category"] != "none"]
    headers = [
        "upvotes", "subreddit", "pain", "goal", "emotion", "life_stage",
        "comment_body", "post_title", "knowledge_gap", "post_url",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    for r, c in enumerate(rated[:1000], start=2):
        write_row(ws, r, [
            c["upvotes"],
            c["subreddit_name"],
            c["pain_category"],
            c["goal_category"],
            c["emotion"],
            c["life_stage"],
            (c["body"] or "")[:2000],
            (c["post_title"] or "")[:120],
            c["knowledge_gap"],
            c["post_url"],
        ], alt=(r % 2 == 0))
    autosize(ws, {
        1: 8, 2: 22, 3: 20, 4: 20, 5: 12, 6: 14, 7: 80, 8: 50, 9: 50, 10: 50,
    })
    ws.freeze_panes = "A2"
    if rated:
        ws.auto_filter.ref = ws.dimensions


def sheet_comments_crosstab(
    wb: Workbook, comments: list[sqlite3.Row], field: str = "pain_category"
) -> None:
    ws = wb.create_sheet("Comments × Subreddit")
    cross: dict[tuple[str, str], int] = Counter()
    field_vals: Counter[str] = Counter()
    sub_vals: Counter[str] = Counter()
    for c in comments:
        cat = c[field]
        if not cat or cat == "none":
            continue
        cross[(c["subreddit_name"], cat)] += 1
        field_vals[cat] += 1
        sub_vals[c["subreddit_name"]] += 1

    if not cross:
        ws["A1"] = "No classified comments yet — run `cli.py classify --target comments`."
        ws["A1"].font = BODY_FONT
        return

    cats = [c for c, _ in field_vals.most_common()]
    subs_sorted = [s for s, _ in sub_vals.most_common()]

    ws.cell(row=1, column=1, value="subreddit \\ comment_pain")
    for i, c in enumerate(cats, start=2):
        ws.cell(row=1, column=i, value=c)
    ws.cell(row=1, column=len(cats) + 2, value="TOTAL")
    style_header(ws, 1, len(cats) + 2)

    vmax = max(cross.values())
    for r, sub in enumerate(subs_sorted, start=2):
        ws.cell(row=r, column=1, value=sub).font = Font(bold=True, name="Helvetica Neue", size=10)
        ws.cell(row=r, column=1).border = BORDER
        for c_idx, c in enumerate(cats, start=2):
            val = cross.get((sub, c), 0)
            cell = ws.cell(row=r, column=c_idx, value=val if val else None)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = BORDER
        ws.cell(row=r, column=len(cats) + 2, value=sub_vals[sub]).font = Font(bold=True)
        ws.cell(row=r, column=len(cats) + 2).border = BORDER
        ws.cell(row=r, column=len(cats) + 2).alignment = CENTER

    total_row = len(subs_sorted) + 2
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=1).fill = H_FILL
    for c_idx, c in enumerate(cats, start=2):
        ws.cell(row=total_row, column=c_idx, value=field_vals[c]).font = Font(bold=True, color=WHITE)
        ws.cell(row=total_row, column=c_idx).fill = H_FILL
        ws.cell(row=total_row, column=c_idx).alignment = CENTER
    ws.cell(row=total_row, column=len(cats) + 2, value=sum(sub_vals.values())).font = Font(bold=True, color=WHITE)
    ws.cell(row=total_row, column=len(cats) + 2).fill = H_FILL
    ws.cell(row=total_row, column=len(cats) + 2).alignment = CENTER

    start_cell = ws.cell(row=2, column=2).coordinate
    end_cell = ws.cell(row=total_row - 1, column=len(cats) + 1).coordinate
    add_color_scale(ws, f"{start_cell}:{end_cell}", vmax)

    widths = {1: 26}
    for i in range(2, len(cats) + 2):
        widths[i] = max(12, len(cats[i - 2]) + 2)
    widths[len(cats) + 2] = 10
    autosize(ws, widths)
    ws.freeze_panes = "B2"


def sheet_conversations(
    wb: Workbook, posts: list[sqlite3.Row], comments: list[sqlite3.Row]
) -> None:
    """Post + its top-3 child comments, joined into a readable block."""
    ws = wb.create_sheet("Top Conversations")
    by_post: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for c in comments:
        by_post[c["post_id"]].append(c)

    # Show the top 30 posts that have any classified comments, sorted by post hotness.
    posts_with_comments = [p for p in posts if by_post.get(p["id"])]
    posts_with_comments.sort(key=lambda p: p["hotness_score"] or 0, reverse=True)
    posts_with_comments = posts_with_comments[:30]

    headers = ["row_type", "upvotes", "subreddit", "pain", "emotion", "text", "post_url"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))

    row = 2
    for p in posts_with_comments:
        # Post row
        write_row(ws, row, [
            "POST",
            p["upvotes"],
            p["subreddit_name"],
            p["pain_category"],
            p["emotion"],
            f"{p['title']}\n\n{(p['selftext'] or '')[:600]}",
            p["url"],
        ])
        # Color-code post row
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=GREY_HEAD)
            ws.cell(row=row, column=col).font = Font(bold=True, name="Helvetica Neue", size=10)
        row += 1
        # Top 3 comments
        children = sorted(
            by_post[p["id"]],
            key=lambda c: c["upvotes"] or 0,
            reverse=True,
        )[:3]
        for c in children:
            write_row(ws, row, [
                f'  ↳ comment d{c["depth"] or 0}',
                c["upvotes"],
                c["subreddit_name"],
                c["pain_category"],
                c["emotion"],
                (c["body"] or "")[:800],
                "",
            ])
            row += 1
        row += 1  # blank spacer
    autosize(ws, {1: 18, 2: 9, 3: 22, 4: 22, 5: 12, 6: 85, 7: 50})


def _safe_split(s: str | None) -> list[str]:
    """Comma-separated text → clean list, tolerant of None/empty."""
    if not s:
        return []
    return [t.strip() for t in s.split(",") if t.strip()]


def sheet_buying_signals(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    """UA segmentation: hot/warm/cold leads based on buying_signal field."""
    ws = wb.create_sheet("Buying Signals")
    by_signal: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for p in posts:
        sig = p["buying_signal"] or "unknown"
        by_signal[sig].append(p)

    ws["A1"] = "UA Lead Segmentation — by buying signal"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE, name="Helvetica Neue")
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28
    ws["A1"].alignment = CENTER

    ws["A3"] = "Tier"
    ws["B3"] = "Signal"
    ws["C3"] = "Posts"
    ws["D3"] = "% of total"
    ws["E3"] = "Top pain"
    ws["F3"] = "Top goal"
    style_header(ws, 3, 6)

    tier_map = {
        "paying_for_solution_now": ("🔥 HOTTEST", RED),
        "actively_looking_for_resource": ("🟠 HOT", AMBER),
        "asking_for_advisor": ("🟠 HOT", AMBER),
        "research_phase": ("🟡 WARM", TEAL),
        "no_buying_signal": ("🔵 COLD", NAVY),
        "unknown": ("⚪ UNKNOWN", "808080"),
    }
    total = len(posts)
    row = 4
    for sig in ["paying_for_solution_now", "actively_looking_for_resource",
                "asking_for_advisor", "research_phase",
                "no_buying_signal", "unknown"]:
        items = by_signal.get(sig, [])
        n = len(items)
        if n == 0:
            continue
        pct = round(100 * n / max(total, 1), 1)
        pains = Counter(p["pain_category"] for p in items if p["pain_category"] not in ("none", None))
        goals = Counter(p["goal_category"] for p in items if p["goal_category"] not in ("none", None))
        tier_label, tier_color = tier_map.get(sig, ("?", "808080"))
        ws.cell(row=row, column=1, value=tier_label)
        ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor=tier_color)
        ws.cell(row=row, column=1).font = Font(bold=True, color=WHITE)
        ws.cell(row=row, column=2, value=sig)
        ws.cell(row=row, column=3, value=n)
        ws.cell(row=row, column=4, value=f"{pct}%")
        ws.cell(row=row, column=5, value=pains.most_common(1)[0][0] if pains else "-")
        ws.cell(row=row, column=6, value=goals.most_common(1)[0][0] if goals else "-")
        for c in range(2, 7):
            ws.cell(row=row, column=c).font = BODY_FONT
            ws.cell(row=row, column=c).border = BORDER
            ws.cell(row=row, column=c).alignment = CENTER
        ws.cell(row=row, column=1).border = BORDER
        row += 1

    # Sample of HOTTEST posts (actively_looking + paying_now)
    ws[f"A{row+2}"] = "Sample HOT leads — posts to study for messaging:"
    ws[f"A{row+2}"].font = Font(bold=True, size=12)
    row += 4
    ws.cell(row=row, column=1, value="upvotes")
    ws.cell(row=row, column=2, value="subreddit")
    ws.cell(row=row, column=3, value="signal")
    ws.cell(row=row, column=4, value="title")
    ws.cell(row=row, column=5, value="intent")
    ws.cell(row=row, column=6, value="url")
    style_header(ws, row, 6)
    row += 1
    hot_posts = (
        by_signal.get("paying_for_solution_now", []) +
        by_signal.get("actively_looking_for_resource", []) +
        by_signal.get("asking_for_advisor", [])
    )
    hot_posts.sort(key=lambda p: p["upvotes"] or 0, reverse=True)
    for p in hot_posts[:40]:
        write_row(ws, row, [
            p["upvotes"], p["subreddit_name"], p["buying_signal"],
            (p["title"] or "")[:90], p["intent"] or "-", p["url"],
        ], alt=(row % 2 == 0))
        row += 1
    autosize(ws, {1: 14, 2: 22, 3: 32, 4: 70, 5: 40, 6: 50})
    ws.freeze_panes = "A4"


def sheet_trigger_events(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    """Frequency of life events motivating posts — drives campaign timing."""
    ws = wb.create_sheet("Trigger Events")
    by_trigger: dict[str, list[sqlite3.Row]] = defaultdict(list)
    for p in posts:
        ev = p["trigger_event"] or "unknown"
        if ev in ("none", "unknown"):
            continue
        by_trigger[ev].append(p)

    ws["A1"] = "Life Triggers → Campaign Timing"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE, name="Helvetica Neue")
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws["A3"] = "trigger_event"
    ws["B3"] = "count"
    ws["C3"] = "top_pain"
    ws["D3"] = "top_goal"
    ws["E3"] = "top_buying_signal"
    ws["F3"] = "sample_title"
    style_header(ws, 3, 6)

    row = 4
    for ev, items in sorted(by_trigger.items(), key=lambda x: -len(x[1])):
        pains = Counter(p["pain_category"] for p in items if p["pain_category"] not in ("none", None))
        goals = Counter(p["goal_category"] for p in items if p["goal_category"] not in ("none", None))
        signals = Counter(p["buying_signal"] for p in items if p["buying_signal"] not in ("unknown", None))
        sample = items[0]
        write_row(ws, row, [
            ev, len(items),
            pains.most_common(1)[0][0] if pains else "-",
            goals.most_common(1)[0][0] if goals else "-",
            signals.most_common(1)[0][0] if signals else "-",
            (sample["title"] or "")[:80],
        ], alt=(row % 2 == 0))
        row += 1
    autosize(ws, {1: 28, 2: 8, 3: 22, 4: 22, 5: 30, 6: 70})
    ws.freeze_panes = "A4"


def sheet_brand_mentions(
    wb: Workbook, posts: list[sqlite3.Row], comments: list[sqlite3.Row]
) -> None:
    """Brands/personalities mentioned — competitive intel + influencer targeting."""
    ws = wb.create_sheet("Brand Mentions")
    brand_count: Counter[str] = Counter()
    brand_pain: dict[str, Counter[str]] = defaultdict(Counter)
    brand_buying: dict[str, Counter[str]] = defaultdict(Counter)

    for row in list(posts) + list(comments or []):
        brands = _safe_split(row["mentioned_brands"])
        pain = row["pain_category"] or "none"
        signal = row["buying_signal"] or "unknown"
        for b in brands:
            brand_count[b] += 1
            brand_pain[b][pain] += 1
            brand_buying[b][signal] += 1

    if not brand_count:
        ws["A1"] = "No brand mentions found yet (run classify with --no-keyword-filter to capture)."
        return

    ws["A1"] = "Brand Mentions — competitors, influencers, tools"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["brand", "mentions", "top_associated_pain", "top_buying_signal", "interpretation"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, 5)

    row = 4
    for brand, n in brand_count.most_common(60):
        top_pain = brand_pain[brand].most_common(1)[0][0] if brand_pain[brand] else "-"
        top_signal = brand_buying[brand].most_common(1)[0][0] if brand_buying[brand] else "-"
        interp = ""
        if top_signal == "paying_for_solution_now":
            interp = "📌 paying customers of this brand"
        elif top_signal == "actively_looking_for_resource":
            interp = "👀 considering this brand"
        write_row(ws, row, [brand, n, top_pain, top_signal, interp], alt=(row % 2 == 0))
        row += 1
    autosize(ws, {1: 28, 2: 10, 3: 22, 4: 30, 5: 40})
    ws.freeze_panes = "A4"


def sheet_personas(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    """Audience personas — life_stage × experience × employment."""
    ws = wb.create_sheet("UA Personas")
    persona: Counter[tuple[str, str, str]] = Counter()
    for p in posts:
        key = (
            p["life_stage"] or "unknown",
            p["financial_experience"] or "unknown",
            p["employment_status"] or "unknown",
        )
        persona[key] += 1

    ws["A1"] = "Audience Personas (Life stage × Experience × Employment)"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:G1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["life_stage", "experience", "employment", "count", "top_pain", "top_goal", "top_trigger"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    row = 4
    for (ls, exp, emp), n in persona.most_common(50):
        bucket = [p for p in posts
                  if (p["life_stage"] or "unknown") == ls
                  and (p["financial_experience"] or "unknown") == exp
                  and (p["employment_status"] or "unknown") == emp]
        pains = Counter(p["pain_category"] for p in bucket if p["pain_category"] not in ("none", None))
        goals = Counter(p["goal_category"] for p in bucket if p["goal_category"] not in ("none", None))
        trigs = Counter(p["trigger_event"] for p in bucket if p["trigger_event"] not in ("none", "unknown", None))
        write_row(ws, row, [
            ls, exp, emp, n,
            pains.most_common(1)[0][0] if pains else "-",
            goals.most_common(1)[0][0] if goals else "-",
            trigs.most_common(1)[0][0] if trigs else "-",
        ], alt=(row % 2 == 0))
        row += 1
    autosize(ws, {1: 18, 2: 14, 3: 22, 4: 8, 5: 22, 6: 22, 7: 22})
    ws.freeze_panes = "A4"


def sheet_interests(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    """Frequency of interest tags + which pain each tag co-occurs with."""
    ws = wb.create_sheet("Interest Map")
    tag_count: Counter[str] = Counter()
    tag_pain: dict[str, Counter[str]] = defaultdict(Counter)
    for p in posts:
        for t in _safe_split(p["interests"]):
            tag_count[t] += 1
            tag_pain[t][p["pain_category"] or "none"] += 1
    if not tag_count:
        ws["A1"] = "No interest tags captured yet."
        return

    ws["A1"] = "Interest Tags Frequency (US finance topics)"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["tag", "mentions", "top_associated_pain", "%_with_pain"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    row = 4
    for tag, n in tag_count.most_common(50):
        top_pain_cat, top_pain_n = tag_pain[tag].most_common(1)[0] if tag_pain[tag] else ("-", 0)
        pct = round(100 * top_pain_n / n, 1) if n else 0
        write_row(ws, row, [tag, n, top_pain_cat, f"{pct}%"], alt=(row % 2 == 0))
        row += 1
    autosize(ws, {1: 24, 2: 10, 3: 24, 4: 14})
    ws.freeze_panes = "A4"


def sheet_intent_library(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    """All the concrete next-actions users want to take."""
    ws = wb.create_sheet("Intent Library")
    by_goal: dict[str, list[tuple[str, sqlite3.Row]]] = defaultdict(list)
    for p in posts:
        if not p["intent"]:
            continue
        by_goal[p["goal_category"] or "none"].append((p["intent"], p))

    ws["A1"] = "Intent Library — concrete next-actions users want to take"
    ws["A1"].font = Font(bold=True, size=14, color=WHITE)
    ws["A1"].fill = H_FILL
    ws.merge_cells("A1:E1")
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    headers = ["goal_category", "intent", "buying_signal", "upvotes", "sample_title"]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=3, column=i, value=h)
    style_header(ws, 3, len(headers))

    row = 4
    for goal in sorted(by_goal.keys(), key=lambda k: -len(by_goal[k])):
        items = by_goal[goal]
        items.sort(key=lambda x: x[1]["upvotes"] or 0, reverse=True)
        for intent, p in items[:30]:
            write_row(ws, row, [
                goal, intent, p["buying_signal"] or "?",
                p["upvotes"], (p["title"] or "")[:80],
            ], alt=(row % 2 == 0))
            row += 1
    autosize(ws, {1: 22, 2: 50, 3: 32, 4: 9, 5: 60})
    ws.freeze_panes = "A4"


def sheet_raw(wb: Workbook, posts: list[sqlite3.Row]) -> None:
    ws = wb.create_sheet("Raw Data")
    headers = [
        "id", "subreddit", "title", "selftext", "upvotes", "num_comments", "hotness",
        "pain", "pain_other", "goal", "goal_other", "knowledge_gap", "emotion",
        "life_stage", "experience", "employment", "family",
        "intent", "interests",
        "trigger", "buying_signal", "urgency", "brands",
        "quotable", "url",
    ]
    for i, h in enumerate(headers, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(headers))
    for r, p in enumerate(posts, start=2):
        write_row(ws, r, [
            p["id"], p["subreddit_name"], p["title"], (p["selftext"] or "")[:3000],
            p["upvotes"], p["num_comments"], p["hotness_score"],
            p["pain_category"], p["pain_other"], p["goal_category"], p["goal_other"],
            p["knowledge_gap"], p["emotion"],
            p["life_stage"], p["financial_experience"], p["employment_status"], p["family_status"],
            p["intent"], p["interests"],
            p["trigger_event"], p["buying_signal"], p["urgency"], p["mentioned_brands"],
            p["quote_worthy"], p["url"],
        ], alt=(r % 2 == 0))
    autosize(ws, {
        1: 14, 2: 22, 3: 50, 4: 60, 5: 8, 6: 8, 7: 9,
        8: 22, 9: 22, 10: 22, 11: 22, 12: 45, 13: 12,
        14: 14, 15: 14, 16: 18, 17: 18,
        18: 40, 19: 50,
        20: 22, 21: 30, 22: 18, 23: 30,
        24: 9, 25: 40,
    })
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# --- Orchestration --------------------------------------------------------


def build(output_path: Path) -> Path:
    with sqlite3.connect(DB_PATH) as conn:
        conn.row_factory = sqlite3.Row
        posts = fetch_all(conn)
        demos = fetch_demographics(conn)
        subs_meta = fetch_subreddits(conn)
        comments = fetch_comments_classified(conn)

    if not posts:
        raise RuntimeError(
            "No classified posts in DB. Run `python cli.py classify` first."
        )

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    sheet_summary(wb, posts, subs_meta, comments)
    # UA-focused (most important for user acquisition research)
    sheet_buying_signals(wb, posts)
    sheet_trigger_events(wb, posts)
    sheet_personas(wb, posts)
    sheet_brand_mentions(wb, posts, comments)
    sheet_interests(wb, posts)
    sheet_intent_library(wb, posts)
    # Classic pain/goal analytics
    sheet_crosstab(wb, "Top Pains", posts, "pain_category")
    sheet_crosstab(wb, "Top Goals", posts, "goal_category")
    sheet_verbatims(wb, posts)
    sheet_gaps(wb, posts)
    sheet_demographics(wb, posts, demos)
    sheet_subreddit_profiles(wb, posts, subs_meta)
    if comments:
        sheet_top_comments(wb, comments)
        sheet_comments_crosstab(wb, comments, "pain_category")
        sheet_conversations(wb, posts, comments)
    sheet_trends(wb, posts)
    sheet_raw(wb, posts)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--output", type=Path,
        default=Path("reports") / f"reddit_research_{dt.date.today().isoformat()}.xlsx",
    )
    ap.add_argument(
        "--upload", action="store_true",
        help="Push xlsx to Google Drive (gdrive:проекты/reddit_parser/)",
    )
    ap.add_argument(
        "--replace", action="store_true",
        help="Overwrite existing file with same name (preserves Drive URL/ID)",
    )
    args = ap.parse_args(argv)

    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    out = build(args.output)
    logger.info("Wrote %s (%.1f KB)", out, out.stat().st_size / 1024)

    if args.upload:
        from drive_upload import push_to_drive  # lazy import
        push_to_drive(out, replace=args.replace)

    print(f"\n✓ Report: {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
